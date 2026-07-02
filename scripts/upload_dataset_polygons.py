#!/usr/bin/env python3
"""
Upload dataset polygon geometry and connect it to USES ties.

The importer expects an outer zip containing:
  - one spreadsheet with USES tie rows
  - one nested zip containing an ESRI shapefile

By default this is a dry run. Pass --apply to write:
  - :GEOMETRY nodes in gisdb
  - USES.geoPolygon links in the target CatMapper database
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from CM.utils import closeAllDrivers, getDriver, getQuery  # noqa: E402


REQUIRED_COLUMNS = ["datasetName", "datasetID", "CMID", "CMName", "Key", "Name", "label"]
FEATURE_ID_PATTERN = re.compile(r"Feature\s+ID\s*==\s*(\d+)", re.IGNORECASE)


@dataclass(frozen=True)
class PreparedData:
    records: List[Dict[str, Any]]
    geojson_path: Path
    workbook_path: Path
    shapefile_path: Path


def load_renviron(path: Path | None = None) -> None:
    renviron = path or Path.home() / ".Renviron"
    if not renviron.exists():
        return

    for raw_line in renviron.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", key):
            continue
        value = value.strip().strip("\"'")
        os.environ.setdefault(key, value)


def reset_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def unzip_archive(zip_path: Path, output_dir: Path) -> None:
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(output_dir)


def find_one(paths: Iterable[Path], suffix: str, label: str) -> Path:
    matches = sorted(path for path in paths if path.suffix.lower() == suffix.lower())
    if len(matches) != 1:
        formatted = ", ".join(str(path) for path in matches) or "none"
        raise ValueError(f"Expected exactly one {label}; found {len(matches)}: {formatted}")
    return matches[0]


def extract_inputs(bundle_zip: Path, work_dir: Path) -> tuple[Path, Path]:
    extract_dir = work_dir / "extract"
    nested_dir = work_dir / "nested"
    reset_dir(extract_dir)
    reset_dir(nested_dir)

    unzip_archive(bundle_zip, extract_dir)
    all_files = [path for path in extract_dir.rglob("*") if path.is_file()]
    workbook = find_one(all_files, ".xlsx", "Excel workbook")
    nested_zip = find_one([path for path in all_files if path != bundle_zip], ".zip", "nested shapefile zip")

    unzip_archive(nested_zip, nested_dir)
    shapefile = find_one([path for path in nested_dir.rglob("*") if path.is_file()], ".shp", "shapefile")
    return workbook, shapefile


def convert_shapefile(shapefile: Path, output_path: Path, simplify_tolerance: float | None) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        output_path.unlink()

    cmd = [
        "ogr2ogr",
        "-f",
        "GeoJSON",
        "-t_srs",
        "EPSG:4326",
        "-preserve_fid",
    ]
    if simplify_tolerance is not None:
        cmd.extend(["-simplify", str(simplify_tolerance)])
    cmd.extend([str(output_path), str(shapefile)])
    subprocess.run(cmd, check=True)
    return output_path


def parse_feature_id(key: Any) -> int:
    match = FEATURE_ID_PATTERN.search(str(key or ""))
    if not match:
        raise ValueError(f"Could not parse Feature ID from Key value: {key!r}")
    return int(match.group(1))


def read_workbook_rows(workbook: Path, dataset_id: str) -> List[Dict[str, Any]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Workbook is empty: {workbook}")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in header]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}")

    index = {column: header.index(column) for column in REQUIRED_COLUMNS}
    parsed = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(value is None for value in row):
            continue
        record = {column: row[index[column]] for column in REQUIRED_COLUMNS}
        record["rowNumber"] = row_number
        record["featureID"] = parse_feature_id(record["Key"])
        if str(record["datasetID"]).strip() != dataset_id:
            raise ValueError(
                f"Row {row_number} has datasetID {record['datasetID']!r}; expected {dataset_id!r}"
            )
        parsed.append(record)

    feature_ids = [row["featureID"] for row in parsed]
    duplicate_ids = sorted({fid for fid in feature_ids if feature_ids.count(fid) > 1})
    if duplicate_ids:
        raise ValueError(f"Duplicate Feature IDs in workbook: {duplicate_ids[:20]}")

    keys = [str(row["Key"]).strip() for row in parsed]
    duplicate_keys = sorted({key for key in keys if keys.count(key) > 1})
    if duplicate_keys:
        raise ValueError(f"Duplicate Key values in workbook: {duplicate_keys[:20]}")

    return parsed


def load_geojson_features(geojson_path: Path) -> Dict[int, Dict[str, Any]]:
    with geojson_path.open(encoding="utf-8") as handle:
        geojson = json.load(handle)

    features = {}
    for fallback_id, feature in enumerate(geojson.get("features") or []):
        raw_id = feature.get("id", fallback_id)
        try:
            feature_id = int(raw_id)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"GeoJSON feature has non-integer id: {raw_id!r}") from exc
        if feature_id in features:
            raise ValueError(f"Duplicate GeoJSON feature id: {feature_id}")
        features[feature_id] = feature

    return features


def compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def prepare_records(
    workbook: Path,
    geojson_path: Path,
    shapefile: Path,
    dataset_id: str,
    geom_prefix: str,
) -> List[Dict[str, Any]]:
    rows = read_workbook_rows(workbook, dataset_id=dataset_id)
    features = load_geojson_features(geojson_path)
    expected_ids = {row["featureID"] for row in rows}
    actual_ids = set(features)

    missing_features = sorted(expected_ids - actual_ids)
    extra_features = sorted(actual_ids - expected_ids)
    if missing_features or extra_features:
        raise ValueError(
            "Workbook Feature IDs and GeoJSON feature IDs differ. "
            f"Missing features: {missing_features[:20]}; extra features: {extra_features[:20]}"
        )

    records = []
    for row in rows:
        feature = features[row["featureID"]]
        properties = feature.get("properties") or {}
        geometry = feature.get("geometry")
        if not geometry:
            raise ValueError(f"Feature {row['featureID']} has no geometry")

        geom_id = f"{geom_prefix}{row['featureID']}"
        records.append(
            {
                "datasetName": str(row["datasetName"]).strip(),
                "datasetID": str(row["datasetID"]).strip(),
                "CMID": str(row["CMID"]).strip(),
                "CMName": str(row["CMName"]).strip(),
                "Key": str(row["Key"]).strip(),
                "usesName": str(row["Name"]).strip(),
                "label": str(row["label"]).strip(),
                "featureID": int(row["featureID"]),
                "geomID": geom_id,
                "origin": f"{dataset_id}:Feature ID == {row['featureID']}",
                "geometry": compact_json(geometry),
                "shapeName": properties.get("Name"),
                "shapeOther": properties.get("Other"),
                "languageFamily": properties.get("Language f"),
                "shapefile": shapefile.name,
            }
        )

    return records


def prepare_data(args: argparse.Namespace) -> PreparedData:
    bundle_zip = Path(args.bundle_zip).resolve()
    work_dir = Path(args.work_dir).resolve()
    if not bundle_zip.exists():
        raise FileNotFoundError(bundle_zip)

    reset_dir(work_dir)
    workbook, shapefile = extract_inputs(bundle_zip, work_dir)
    geojson_path = convert_shapefile(
        shapefile=shapefile,
        output_path=work_dir / "converted.geojson",
        simplify_tolerance=args.simplify_tolerance,
    )
    geom_prefix = args.geom_prefix or f"{args.dataset_id}_Feature_ID_"
    records = prepare_records(
        workbook=workbook,
        geojson_path=geojson_path,
        shapefile=shapefile,
        dataset_id=args.dataset_id,
        geom_prefix=geom_prefix,
    )
    return PreparedData(
        records=records,
        geojson_path=geojson_path,
        workbook_path=workbook,
        shapefile_path=shapefile,
    )


def preflight_target_database(database: str, records: List[Dict[str, Any]]) -> Dict[str, Any]:
    driver = getDriver(database)

    property_rows = getQuery(
        """
        OPTIONAL MATCH (p:PROPERTY {CMName: 'geoPolygon'})
        RETURN p.CMName AS property, p.type AS type
        """,
        driver=driver,
        type="dict",
    )
    has_geo_property = bool(property_rows and property_rows[0].get("property"))

    rows = getQuery(
        """
        UNWIND $rows AS row
        OPTIONAL MATCH (d:DATASET {CMID: row.datasetID})
        OPTIONAL MATCH (c:CATEGORY {CMID: row.CMID})
        OPTIONAL MATCH (d)-[r:USES {Key: row.Key}]->(existing:CATEGORY)
        WITH row, d, c,
             collect(
                CASE
                    WHEN existing IS NULL THEN NULL
                    ELSE {
                        CMID: existing.CMID,
                        CMName: existing.CMName,
                        geoPolygon: r.geoPolygon,
                        relID: elementId(r)
                    }
                END
             ) AS rawRels
        RETURN row.featureID AS featureID,
               row.Key AS Key,
               row.CMID AS expectedCMID,
               row.geomID AS expectedGeomID,
               d IS NOT NULL AS datasetExists,
               c IS NOT NULL AS categoryExists,
               [rel IN rawRels WHERE rel IS NOT NULL] AS existingRels
        ORDER BY featureID
        """,
        driver=driver,
        params={"rows": records},
        type="dict",
    )

    missing_categories = [
        {"featureID": row["featureID"], "CMID": row["expectedCMID"], "Key": row["Key"]}
        for row in rows
        if not row.get("categoryExists")
    ]
    dataset_missing = [row for row in rows if not row.get("datasetExists")]
    conflicts = []
    existing = 0
    existing_with_geo = 0
    missing_uses = 0

    for row in rows:
        rels = row.get("existingRels") or []
        if not rels:
            missing_uses += 1
            continue
        expected = row["expectedCMID"]
        wrong_targets = [rel for rel in rels if rel.get("CMID") != expected]
        if wrong_targets:
            conflicts.append({"featureID": row["featureID"], "Key": row["Key"], "existing": wrong_targets})
        expected_target_rels = [rel for rel in rels if rel.get("CMID") == expected]
        existing += len(expected_target_rels)
        existing_with_geo += sum(1 for rel in expected_target_rels if rel.get("geoPolygon"))
        if len(expected_target_rels) > 1:
            conflicts.append({"featureID": row["featureID"], "Key": row["Key"], "existing": expected_target_rels})

    return {
        "hasGeoPolygonProperty": has_geo_property,
        "datasetMissingCount": len(dataset_missing),
        "missingCategories": missing_categories,
        "existingUses": existing,
        "existingUsesWithGeoPolygon": existing_with_geo,
        "missingUses": missing_uses,
        "conflicts": conflicts,
    }


def preflight_gis_database(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    driver = getDriver("gisdb")
    rows = getQuery(
        """
        UNWIND $rows AS row
        OPTIONAL MATCH (g:GEOMETRY {geomID: row.geomID})
        OPTIONAL MATCH (origin:GEOMETRY {origin: row.origin})
        WITH row, g, collect(origin.geomID) AS originGeomIDs
        RETURN row.featureID AS featureID,
               row.geomID AS geomID,
               g IS NOT NULL AS geometryExists,
               [geomID IN originGeomIDs WHERE geomID IS NOT NULL AND geomID <> row.geomID] AS originCollisions
        ORDER BY featureID
        """,
        driver=driver,
        params={"rows": records},
        type="dict",
    )

    existing = [row["geomID"] for row in rows if row.get("geometryExists")]
    collisions = [row for row in rows if row.get("originCollisions")]
    return {
        "existingGeometryCount": len(existing),
        "originCollisions": collisions,
    }


def apply_gis(records: List[Dict[str, Any]]) -> int:
    driver = getDriver("gisdb")
    result = getQuery(
        """
        UNWIND $rows AS row
        MERGE (g:GEOMETRY {geomID: row.geomID})
        ON CREATE SET g.createdAt = toString(datetime())
        SET g.geometry = row.geometry,
            g.origin = row.origin,
            g.sourceDatasetID = row.datasetID,
            g.sourceDatasetName = row.datasetName,
            g.featureID = row.featureID,
            g.Name = row.shapeName,
            g.Other = row.shapeOther,
            g.languageFamily = row.languageFamily,
            g.shapefile = row.shapefile,
            g.crs = 'EPSG:4326',
            g.updatedAt = toString(datetime())
        RETURN count(g) AS updated
        """,
        driver=driver,
        params={"rows": records},
        type="dict",
    )
    return int(result[0]["updated"] if result else 0)


def apply_uses(database: str, records: List[Dict[str, Any]]) -> int:
    driver = getDriver(database)
    result = getQuery(
        """
        UNWIND $rows AS row
        MATCH (d:DATASET {CMID: row.datasetID})
        MATCH (c:CATEGORY {CMID: row.CMID})
        MATCH (d)-[r:USES {Key: row.Key}]->(c)
        SET r.status = coalesce(r.status, 'update'),
            r.geoPolygon = row.geomID
        RETURN count(r) AS updated
        """,
        driver=driver,
        params={"rows": records},
        type="dict",
    )
    return int(result[0]["updated"] if result else 0)


def verify_upload(database: str, records: List[Dict[str, Any]]) -> Dict[str, Any]:
    target_driver = getDriver(database)
    gis_driver = getDriver("gisdb")

    linked = getQuery(
        """
        UNWIND $rows AS row
        MATCH (:DATASET {CMID: row.datasetID})-[r:USES {Key: row.Key}]->(:CATEGORY {CMID: row.CMID})
        WHERE r.geoPolygon = row.geomID
        RETURN count(r) AS linked
        """,
        driver=target_driver,
        params={"rows": records},
        type="dict",
    )
    geometries = getQuery(
        """
        UNWIND $rows AS row
        MATCH (g:GEOMETRY {geomID: row.geomID})
        WHERE g.geometry IS NOT NULL
        RETURN count(g) AS geometries
        """,
        driver=gis_driver,
        params={"rows": records},
        type="dict",
    )

    return {
        "linkedUses": int(linked[0]["linked"] if linked else 0),
        "geometryNodes": int(geometries[0]["geometries"] if geometries else 0),
    }


def abort_on_preflight_errors(target: Dict[str, Any], gis: Dict[str, Any]) -> None:
    errors = []
    if not target["hasGeoPolygonProperty"]:
        errors.append("target database is missing relationship PROPERTY metadata for geoPolygon")
    if target["datasetMissingCount"]:
        errors.append(f"dataset is missing in target database ({target['datasetMissingCount']} row checks)")
    if target["missingCategories"]:
        sample = target["missingCategories"][:10]
        errors.append(f"missing CATEGORY nodes: {sample}")
    if target["conflicts"]:
        sample = target["conflicts"][:10]
        errors.append(f"existing USES ties conflict with spreadsheet targets: {sample}")
    if target["missingUses"]:
        errors.append(
            f"{target['missingUses']} USES ties are missing; create them with the standard upload flow first"
        )
    if gis["originCollisions"]:
        sample = gis["originCollisions"][:10]
        errors.append(f"GIS origin collisions: {sample}")

    if errors:
        joined = "\n  - ".join(errors)
        raise RuntimeError(f"Preflight failed:\n  - {joined}")


def print_summary(args: argparse.Namespace, prepared: PreparedData, target: Dict[str, Any], gis: Dict[str, Any]) -> None:
    print(f"Bundle: {Path(args.bundle_zip).resolve()}")
    print(f"Workbook: {prepared.workbook_path}")
    print(f"Shapefile: {prepared.shapefile_path}")
    print(f"Converted GeoJSON: {prepared.geojson_path}")
    print(f"Target database: {args.database}")
    print(f"Dataset CMID: {args.dataset_id}")
    print(f"Prepared records: {len(prepared.records)}")
    print(f"Existing USES ties for these keys: {target['existingUses']}")
    print(f"Existing USES ties already carrying geoPolygon: {target['existingUsesWithGeoPolygon']}")
    print(f"Missing USES ties: {target['missingUses']}")
    print(f"Existing GEOMETRY nodes for these geomIDs: {gis['existingGeometryCount']}")
    if prepared.records:
        sample = prepared.records[0]
        print(
            "Sample mapping: "
            f"{sample['Key']} -> {sample['CMID']} -> {sample['geomID']} "
            f"({sample['shapeName']})"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Upload polygon geometry from a shapefile bundle and attach it to dataset USES ties."
    )
    parser.add_argument("--bundle-zip", default="/mnt/storage/app/tmp/Perreault.zip")
    parser.add_argument("--work-dir", default="/mnt/storage/app/tmp/perreault_polygon_upload")
    parser.add_argument("--database", default="sociomap", choices=["sociomap", "archamap"])
    parser.add_argument("--dataset-id", default="SD2235")
    parser.add_argument("--geom-prefix", default=None)
    parser.add_argument(
        "--simplify-tolerance",
        type=float,
        default=None,
        help="Optional ogr2ogr -simplify tolerance in EPSG:4326 degrees.",
    )
    parser.add_argument("--apply", action="store_true", help="Write GEOMETRY nodes and USES.geoPolygon links.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    load_renviron()

    prepared = prepare_data(args)
    target = preflight_target_database(args.database, prepared.records)
    gis = preflight_gis_database(prepared.records)

    print_summary(args, prepared, target, gis)
    abort_on_preflight_errors(target, gis)

    if not args.apply:
        print("Dry run complete. Re-run with --apply to write changes.")
        return 0

    gis_updated = apply_gis(prepared.records)
    uses_updated = apply_uses(args.database, prepared.records)
    verification = verify_upload(args.database, prepared.records)

    print(f"GEOMETRY nodes written: {gis_updated}")
    print(f"USES ties written: {uses_updated}")
    print(f"Verified GEOMETRY nodes: {verification['geometryNodes']}")
    print(f"Verified USES.geoPolygon links: {verification['linkedUses']}")

    if verification["geometryNodes"] != len(prepared.records) or verification["linkedUses"] != len(prepared.records):
        raise RuntimeError(f"Post-upload verification did not match prepared record count ({len(prepared.records)}).")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    finally:
        closeAllDrivers()
